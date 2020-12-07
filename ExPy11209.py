# 파이썬 기초부터 활용까지 (2020.09)
# [12과] 엑셀 작업

from openpyxl import *

wb = load_workbook('ExPy11203.xlsx')
ws = wb.active

ws.insert_rows(5)
ws.insert_cols(2, 3)
ws.delete_rows(2, 2)
ws.delete_rows(5, 2)

ws.merge_cells('A1:B2')

ws1 = wb.create_sheet('newsheet1')
ws2 = wb.create_sheet('newsheet2', 0)

wb.save('ExPy11209.xlsx')