# 파이썬 기초부터 활용까지 (2020.09)
# [12과] 엑셀 작업

# column과 row의 선택적 작업
from openpyxl import *
from openpyxl.styles import *

wb = Workbook()
ws = wb.active

for i in range(1, 10):
    for j in range(1, 10):
        ws.cell(i, j, i*j)

c = ws['C']
for col in c:
    col.font = Font(name='HY헤드라인M', bold=True)
    col.alignment = Alignment(horizontal='center', vertical='center')
    th = Side(border_style='thin')
    col.border = Border(top=th, bottom=th, left=th, right=th)

wb.save('ExPy11207.xlsx')
