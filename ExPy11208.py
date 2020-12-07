# 파이썬 기초부터 활용까지 (2020.09)
# [12과] 엑셀 작업

from openpyxl import *

wb = load_workbook('ExPy11203.xlsx')
ws = wb.active

# 열단위 슬라이싱
col_range = ws['B:C']

for i in col_range:
    for j in i:
        print(j.value)

# 행단위 슬라이싱
row_range = ws[6:7]

for i in row_range:
    for j in i:
        print(j.value)

# 값만 활용하는 작업

for i in ws.values:
    for j in i:
        print(j)