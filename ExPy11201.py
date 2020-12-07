# 파이썬 기초부터 활용까지 (2020.09)
# [12과] 엑셀 작업

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws['a1'] = '안녕하세요'
ws.cell(2,1).value = '파이선'
ws.cell(3,1, '개나리')
wb.save('ExPy11201.xlsx')
