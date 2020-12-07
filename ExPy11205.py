# 파이썬 기초부터 활용까지 (2020.09)
# [12과] 엑셀 작업

# 정렬방식 / 폰트지정 / 선긋기 : styles 객체의 Alignment, Font, Border, Side 클래스 이용

import openpyxl
# from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles import *


wb = openpyxl.Workbook()
ws = wb.active

ws['B3'] = 'Hello'
ws['B3'].font = Font(name = 'HY헤드라인M',
                     bold = True,
                     size = 20,
                     italic=True,
                     underline='single')
ws['B3'].alignment = Alignment(horizontal='center',
                               vertical='center')
th = Side(border_style='thin')
db = Side(border_style='double')
ws['B3'].border = Border(top=th, bottom=th, left=db, right=db)

wb.save('ExPy11205.xlsx')


